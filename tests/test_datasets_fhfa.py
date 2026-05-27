"""Tests for datasets.fhfa loader."""

import os
import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook

from py_tools.datasets import fhfa


def _write_zip3_quarterly_xlsx(path, rows):
    """Write a minimal quarterly zip3 xlsx fixture matching FHFA's layout."""
    wb = Workbook()
    ws = wb.active
    ws.append(["HPI for Three-Digit ZIP Codes (All-Transactions Index)"])
    ws.append([None])
    ws.append(["Disclaimer text placeholder."])
    ws.append(["Last updated: January 1, 2025.", None, None, None, None])
    ws.append([
        "Three-Digit ZIP Code",
        "Year",
        "Quarter",
        "Index (NSA)",
        "Index Type",
    ])
    for row in rows:
        ws.append(row)
    wb.save(path)


def _write_zip3_annual_xlsx(path, rows):
    """Write a minimal annual zip3 xlsx fixture matching FHFA's layout."""
    wb = Workbook()
    ws = wb.active
    ws.append(["HPI for Three-Digit ZIP Codes (All-Transactions Index)"])
    ws.append([None])
    ws.append(["Disclaimer text placeholder."])
    ws.append(["Last updated: January 1, 2025.", None, None, None, None, None, None])
    ws.append(["Not Seasonally Adjusted (NSA) ", None, None, None, None, None])
    ws.append(["Three-Digit ZIP Code", "Year", "Annual Change (%)", "HPI",
               "HPI with 1990 base", "HPI with 2000 base"])
    for row in rows:
        ws.append(row)
    wb.save(path)


def _write_zip5_xlsx(path, rows):
    """Write a minimal zip5 xlsx fixture matching FHFA's file layout."""
    wb = Workbook()
    ws = wb.active
    ws.append(["HPI for Five-Digit ZIP Codes (All-Transactions Index)"])
    ws.append([None])
    ws.append(["Disclaimer text placeholder."])
    ws.append(["Last updated: January 1, 2025.", None, None, None, None, None])
    ws.append(["Not Seasonally Adjusted (NSA) ", None, None, None, None, None])
    ws.append(["Five-Digit ZIP Code", "Year", "Annual Change (%)", "HPI",
               "HPI with 1990 base", "HPI with 2000 base"])
    for row in rows:
        ws.append(row)
    wb.save(path)


@pytest.fixture()
def zip5_dir(tmp_path):
    rows = [
        [10001, 2010, None,  100.0, 80.0, 75.0],
        [10001, 2011,  5.0,  105.0, 84.0, 78.75],
        [10001, 2012,  3.5,  108.68, 86.94, 81.51],
        [10002, 2010, None,  100.0, 90.0, 85.0],
        [10002, 2011,  3.0,  103.0, 92.7, 87.55],
    ]
    _write_zip5_xlsx(tmp_path / "hpi_at_zip5.xlsx", rows)
    return str(tmp_path) + "/"


@pytest.fixture()
def zip3_dir(tmp_path):
    quarterly_rows = [
        [100, 2010, 1, 100.0, "Native 3-Digit ZIP index"],
        [100, 2010, 2, 101.5, "Native 3-Digit ZIP index"],
        [101, 2010, 1, 98.0, "Native 3-Digit ZIP index"],
    ]
    annual_rows = [
        [100, 2010, None, 100.25, 80.0, 75.0],
        [100, 2011,  5.0, 105.5, 84.0, 78.75],
        [101, 2010, None, 100.0, 90.0, 85.0],
    ]
    _write_zip3_quarterly_xlsx(tmp_path / "HPI_AT_3zip.xlsx", quarterly_rows)
    _write_zip3_annual_xlsx(tmp_path / "hpi_at_zip3_annual.xlsx", annual_rows)
    return str(tmp_path) + "/"


def test_zip3_default_is_quarterly(zip3_dir):
    df = fhfa.load("zip3", data_dir=zip3_dir, reimport=True)
    assert df.index.names == ["zip3", "date"]
    assert set(df.columns) == {"year", "quarter", "hpi"}
    assert df.loc[(100, pd.Timestamp("2010-04-01")), "hpi"] == pytest.approx(101.5)


def test_zip3_annual_index_columns_and_values(zip3_dir):
    df = fhfa.load("zip3", annual=True, data_dir=zip3_dir, reimport=True)
    assert df.index.names == ["zip3", "date"]
    assert set(df.columns) >= {"hpi", "hpi_1990_base", "hpi_2000_base", "annual_change_pct"}
    assert isinstance(df.index.get_level_values("date"), pd.DatetimeIndex)
    assert df["hpi"].dtype == np.float64
    assert df["annual_change_pct"].dtype == np.float64
    assert df.loc[(100, pd.Timestamp("2011-01-01")), "hpi"] == pytest.approx(105.5)
    assert np.isnan(df.loc[(100, pd.Timestamp("2010-01-01")), "annual_change_pct"])


def test_zip3_annual_has_separate_parquet_cache(zip3_dir):
    fhfa.load("zip3", data_dir=zip3_dir, reimport=True)
    fhfa.load("zip3", annual=True, data_dir=zip3_dir, reimport=True)

    assert os.path.exists(zip3_dir + "fhfazip3_at.parquet")
    assert os.path.exists(zip3_dir + "fhfazip3_at_annual.parquet")

    df_cached = fhfa.load("zip3", annual=True, data_dir=zip3_dir)
    df_fresh = fhfa.load("zip3", annual=True, data_dir=zip3_dir, reimport=True)
    pd.testing.assert_frame_equal(df_cached, df_fresh)


@pytest.mark.parametrize(
    "kwargs",
    [
        {"dataset": "zip5", "annual": False},
        {"dataset": "county", "annual": False},
        {"dataset": "state", "annual": True},
        {"dataset": "metro", "annual": True},
    ],
)
def test_unsupported_annual_combinations_raise(zip3_dir, kwargs):
    with pytest.raises(ValueError):
        fhfa.load(data_dir=zip3_dir, **kwargs)


def test_zip5_index(zip5_dir):
    df = fhfa.load("zip5", data_dir=zip5_dir, reimport=True)
    assert df.index.names == ["zip5", "date"]


def test_zip5_columns(zip5_dir):
    df = fhfa.load("zip5", data_dir=zip5_dir, reimport=True)
    assert set(df.columns) >= {"hpi", "hpi_1990_base", "hpi_2000_base", "annual_change_pct"}


def test_zip5_dtypes(zip5_dir):
    df = fhfa.load("zip5", data_dir=zip5_dir, reimport=True)
    assert df["hpi"].dtype == np.float64
    assert df["annual_change_pct"].dtype == np.float64
    assert isinstance(df.index.get_level_values("date"), pd.DatetimeIndex)


def test_zip5_values(zip5_dir):
    df = fhfa.load("zip5", data_dir=zip5_dir, reimport=True)
    assert df.loc[(10001, pd.Timestamp("2011-01-01")), "hpi"] == pytest.approx(105.0)
    assert df.loc[(10002, pd.Timestamp("2010-01-01")), "hpi_1990_base"] == pytest.approx(90.0)
    assert np.isnan(df.loc[(10001, pd.Timestamp("2010-01-01")), "annual_change_pct"])


def test_zip5_shape(zip5_dir):
    df = fhfa.load("zip5", data_dir=zip5_dir, reimport=True)
    assert df.shape[0] == 5
    assert df.index.is_unique


def test_zip5_parquet_cache(zip5_dir):
    fhfa.load("zip5", data_dir=zip5_dir, reimport=True)
    parquet_path = zip5_dir + "fhfazip5_at.parquet"
    assert os.path.exists(parquet_path)

    df_cached = fhfa.load("zip5", data_dir=zip5_dir)
    df_fresh = fhfa.load("zip5", data_dir=zip5_dir, reimport=True)
    pd.testing.assert_frame_equal(df_cached, df_fresh)
